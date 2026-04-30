# -*- coding: utf-8 -*-
from __future__ import annotations

import io

from fastapi import HTTPException


def extract_xlsx_text(raw_bytes: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Excel support requires the openpyxl package.") from exc

    workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    rows: list[str] = []
    for sheet in workbook.worksheets:
        rows.append(f"# {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                rows.append(" | ".join(values))
    return "\n".join(rows).strip()


def extract_xls_text(raw_bytes: bytes) -> str:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Legacy Excel support requires the xlrd package.") from exc

    workbook = xlrd.open_workbook(file_contents=raw_bytes)
    rows: list[str] = []
    for sheet in workbook.sheets():
        rows.append(f"# {sheet.name}")
        for row_index in range(sheet.nrows):
            values = [str(cell).strip() for cell in sheet.row_values(row_index) if str(cell).strip()]
            if values:
                rows.append(" | ".join(values))
    return "\n".join(rows).strip()