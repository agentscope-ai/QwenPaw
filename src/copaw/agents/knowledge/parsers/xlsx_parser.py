# -*- coding: utf-8 -*-
"""XLSX parser for knowledge import."""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path

from ..exceptions import EmptyParsedContentError, KnowledgeError
from ..models import ParsedDocument


class XlsxParser:
    """Parser for XLSX workbook files."""

    supported_suffixes = (".xlsx",)

    def parse(self, path: Path) -> ParsedDocument:
        try:
            from openpyxl import (
                load_workbook,
            )  # lazy import to avoid hard import cost
        except ImportError as exc:
            raise KnowledgeError(
                "openpyxl is required for XLSX knowledge import support",
            ) from exc

        workbook = load_workbook(
            filename=str(path),
            read_only=True,
            data_only=False,
        )
        try:
            blocks: list[str] = []
            nonempty_row_count = 0
            nonempty_cell_count = 0

            for sheet in workbook.worksheets:
                for row_index, row in enumerate(
                    sheet.iter_rows(values_only=True),
                    start=1,
                ):
                    values = [_stringify_cell(value) for value in row]
                    while values and not values[-1]:
                        values.pop()
                    if not values or not any(values):
                        continue

                    nonempty_row_count += 1
                    nonempty_cell_count += sum(1 for value in values if value)
                    row_line = " | ".join(values)
                    marker = f"<<<SHEET:{sheet.title} ROW:{row_index}>>>"
                    blocks.append(
                        f"{marker}\n{row_line}",
                    )

            raw_text = "\n\n".join(blocks).strip()
            if not raw_text:
                raise EmptyParsedContentError(
                    "XLSX contains no extractable text content",
                )

            return ParsedDocument(
                title=path.stem,
                source_path=str(path),
                source_type="xlsx",
                raw_text=raw_text,
                metadata={
                    "sheet_count": len(workbook.worksheets),
                    "row_count": nonempty_row_count,
                    "nonempty_cell_count": nonempty_cell_count,
                    "line_count": len(raw_text.splitlines()),
                },
            )
        finally:
            workbook.close()


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
